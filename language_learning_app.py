##
import random
from openpyxl import load_workbook
import tkinter as tk
from tkinter import font


class LanguageLearningApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Language Learning App")

        label_font = font.Font(size=14)
        input_font = font.Font(size=12)

        self.english_label = tk.Label(self, text="English Word:", font=label_font)
        self.english_label.pack()

        self.hint_label = tk.Label(self, text="Hint: ", font=label_font)
        self.hint_label.pack()

        self.german_input = tk.Entry(self, font=input_font)
        self.german_input.pack()

        self.check_button = tk.Button(self, text="Check", command=self.check_translation, font=label_font)
        self.check_button.pack()

        self.result_label = tk.Label(self, text="", font=label_font)
        self.result_label.pack()

        self.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.load_excel_file("a2.xlsx") # df_deleted_learned

    def load_excel_file(self, file_path):
        self.file_path = file_path
        self.words = []
        self.learned_words = []

        workbook = load_workbook(file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            german, english = row
            self.words.append((german, english))

        self.show_random_word()

    def show_random_word(self):
        self.current_word = random.choice(self.words)
        self.english_label.config(text=f"English Word: {self.current_word[1]}")
        self.hint_label.config(text=f"Hint: {self.current_word[0][0]}")

    def check_translation(self):
        user_input = self.german_input.get().strip()
        if user_input == self.current_word[0]:
            self.result_label.config(text="Correct!")
            self.words.remove(self.current_word)
            self.learned_words.append(self.current_word)
        else:
            self.result_label.config(text=f"Incorrect! The correct translation is: {self.current_word[0]}")

        self.show_random_word()
        self.german_input.delete(0, tk.END)

    def update_excel_file(self):
        workbook = load_workbook(self.file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.value = None

        for index, word in enumerate(self.words, start=2):
            sheet.cell(row=index, column=1, value=word[0])
            sheet.cell(row=index, column=2, value=word[1])

        workbook.save(self.file_path)

    def on_closing(self):
        self.update_excel_file()
        self.destroy()


if __name__ == "__main__":
    app = LanguageLearningApp()
    app.mainloop()
